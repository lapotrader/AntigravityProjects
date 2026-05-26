import pandas as pd
import numpy as np
import os

def load_data(filepath):
    """Carica e pulisce i dati BTP dal file di testo daily o intraday."""
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Il file '{filepath}' non esiste.")
        
    df = pd.read_csv(filepath, sep='\t')
    df.columns = [col.strip().lower() for col in df.columns]
    
    # Rilevamento automatico formato data
    df['data'] = pd.to_datetime(df['data'], errors='coerce')
    if df['data'].isna().any():
        df['data'] = pd.to_datetime(df['data'], format='%d/%m/%Y', errors='coerce')
        
    df = df.sort_values('data').reset_index(drop=True)
    
    numeric_cols = ['high', 'low', 'open', 'close', 'volume']
    for col in numeric_cols:
        if col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].astype(str).str.replace(',', '.').astype(float)
            else:
                df[col] = df[col].astype(float)
                
    return df

def calculate_supertrend(df, period=20, multiplier=1.5, atr_type='rma'):
    """Calcola l'indicatore Supertrend e l'ATR (rma = Wilder)."""
    high = df['high']
    low = df['low']
    close = df['close']
    
    tr1 = high - low
    tr2 = (high - close.shift(1)).abs()
    tr3 = (low - close.shift(1)).abs()
    df['tr'] = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
    
    if atr_type == 'sma':
        df['atr'] = df['tr'].rolling(window=period).mean()
    elif atr_type == 'ema':
        df['atr'] = df['tr'].ewm(span=period, adjust=False).mean()
    else: # rma
        df['atr'] = df['tr'].ewm(alpha=1/period, adjust=False).mean()
        
    df['hl2'] = (high + low) / 2
    df['basic_ub'] = df['hl2'] + multiplier * df['atr']
    df['basic_lb'] = df['hl2'] - multiplier * df['atr']
    
    final_ub = [0.0] * len(df)
    final_lb = [0.0] * len(df)
    supertrend = [0.0] * len(df)
    direction = [1] * len(df)
    
    for i in range(len(df)):
        if i == 0:
            final_ub[i] = df['basic_ub'].iloc[i]
            final_lb[i] = df['basic_lb'].iloc[i]
            supertrend[i] = final_ub[i]
            direction[i] = -1
            continue
            
        prev_close = df['close'].iloc[i-1]
        prev_final_ub = final_ub[i-1]
        prev_final_lb = final_lb[i-1]
        
        if df['basic_ub'].iloc[i] < prev_final_ub or prev_close > prev_final_ub:
            final_ub[i] = df['basic_ub'].iloc[i]
        else:
            final_ub[i] = prev_final_ub
            
        if df['basic_lb'].iloc[i] > prev_final_lb or prev_close < prev_final_lb:
            final_lb[i] = df['basic_lb'].iloc[i]
        else:
            final_lb[i] = prev_final_lb
            
        prev_supertrend = supertrend[i-1]
        if prev_supertrend == prev_final_ub:
            if df['close'].iloc[i] > final_ub[i]:
                supertrend[i] = final_lb[i]
                direction[i] = 1
            else:
                supertrend[i] = final_ub[i]
                direction[i] = -1
        else:
            if df['close'].iloc[i] < final_lb[i]:
                supertrend[i] = final_ub[i]
                direction[i] = -1
            else:
                supertrend[i] = final_lb[i]
                direction[i] = 1
                
    df['supertrend'] = supertrend
    df['direction'] = direction
    return df

def run_backtest(df, tp_multiplier=999.0, trade_type='both', ma_col=None):
    """
    Esegue il backtest della strategia Supertrend e raccoglie i trades
    con i dati di mercato (OHLC) del momento di ingresso e di uscita.
    """
    trades = []
    in_position = False
    pos_type = None
    entry_price = 0.0
    entry_idx = 0
    tp_level = 0.0
    
    for i in range(len(df)):
        if i < 21: # Periodo di riscaldamento
            continue
            
        prev_dir = df['direction'].iloc[i-1]
        prev_prev_dir = df['direction'].iloc[i-2] if i >= 2 else prev_dir
        
        current_date = df['data'].iloc[i]
        current_open = df['open'].iloc[i]
        current_high = df['high'].iloc[i]
        current_low = df['low'].iloc[i]
        current_close = df['close'].iloc[i]
        current_supertrend = df['supertrend'].iloc[i]
        
        # 1. Gestione posizione aperta
        if in_position:
            if pos_type == 'long':
                # Verifica Take Profit (intraday)
                if current_high >= tp_level:
                    pnl = tp_level - entry_price
                    trades.append({
                        'tipo': 'Long',
                        'data_ingresso': df['data'].iloc[entry_idx],
                        'prezzo_ingresso': entry_price,
                        'data_uscita': current_date,
                        'prezzo_uscita': tp_level,
                        'motivo_uscita': 'Take Profit',
                        'pnl_punti': pnl,
                        'pnl_eur': pnl * 1000,
                        'durata_barre': i - entry_idx,
                        'mercato_open_ingresso': df['open'].iloc[entry_idx],
                        'mercato_high_ingresso': df['high'].iloc[entry_idx],
                        'mercato_low_ingresso': df['low'].iloc[entry_idx],
                        'mercato_close_ingresso': df['close'].iloc[entry_idx],
                        'mercato_open_uscita': current_open,
                        'mercato_high_uscita': current_high,
                        'mercato_low_uscita': current_low,
                        'mercato_close_uscita': current_close
                    })
                    in_position = False
                    pos_type = None
                # Verifica Stop Loss (in chiusura)
                elif current_close < current_supertrend:
                    pnl = current_close - entry_price
                    trades.append({
                        'tipo': 'Long',
                        'data_ingresso': df['data'].iloc[entry_idx],
                        'prezzo_ingresso': entry_price,
                        'data_uscita': current_date,
                        'prezzo_uscita': current_close,
                        'motivo_uscita': 'Supertrend Stop',
                        'pnl_punti': pnl,
                        'pnl_eur': pnl * 1000,
                        'durata_barre': i - entry_idx,
                        'mercato_open_ingresso': df['open'].iloc[entry_idx],
                        'mercato_high_ingresso': df['high'].iloc[entry_idx],
                        'mercato_low_ingresso': df['low'].iloc[entry_idx],
                        'mercato_close_ingresso': df['close'].iloc[entry_idx],
                        'mercato_open_uscita': current_open,
                        'mercato_high_uscita': current_high,
                        'mercato_low_uscita': current_low,
                        'mercato_close_uscita': current_close
                    })
                    in_position = False
                    pos_type = None
                    
            elif pos_type == 'short':
                # Verifica Take Profit (intraday)
                if current_low <= tp_level:
                    pnl = entry_price - tp_level
                    trades.append({
                        'tipo': 'Short',
                        'data_ingresso': df['data'].iloc[entry_idx],
                        'prezzo_ingresso': entry_price,
                        'data_uscita': current_date,
                        'prezzo_uscita': tp_level,
                        'motivo_uscita': 'Take Profit',
                        'pnl_punti': pnl,
                        'pnl_eur': pnl * 1000,
                        'durata_barre': i - entry_idx,
                        'mercato_open_ingresso': df['open'].iloc[entry_idx],
                        'mercato_high_ingresso': df['high'].iloc[entry_idx],
                        'mercato_low_ingresso': df['low'].iloc[entry_idx],
                        'mercato_close_ingresso': df['close'].iloc[entry_idx],
                        'mercato_open_uscita': current_open,
                        'mercato_high_uscita': current_high,
                        'mercato_low_uscita': current_low,
                        'mercato_close_uscita': current_close
                    })
                    in_position = False
                    pos_type = None
                # Verifica Stop Loss (in chiusura)
                elif current_close > current_supertrend:
                    pnl = entry_price - current_close
                    trades.append({
                        'tipo': 'Short',
                        'data_ingresso': df['data'].iloc[entry_idx],
                        'prezzo_ingresso': entry_price,
                        'data_uscita': current_date,
                        'prezzo_uscita': current_close,
                        'motivo_uscita': 'Supertrend Stop',
                        'pnl_punti': pnl,
                        'pnl_eur': pnl * 1000,
                        'durata_barre': i - entry_idx,
                        'mercato_open_ingresso': df['open'].iloc[entry_idx],
                        'mercato_high_ingresso': df['high'].iloc[entry_idx],
                        'mercato_low_ingresso': df['low'].iloc[entry_idx],
                        'mercato_close_ingresso': df['close'].iloc[entry_idx],
                        'mercato_open_uscita': current_open,
                        'mercato_high_uscita': current_high,
                        'mercato_low_uscita': current_low,
                        'mercato_close_uscita': current_close
                    })
                    in_position = False
                    pos_type = None
                    
        # 2. Nuovi Ingressi (solo se non siamo in posizione)
        if not in_position:
            close_prev = df['close'].iloc[i-1]
            ma_val_prev = df[ma_col].iloc[i-1] if ma_col is not None and ma_col in df.columns else None
            
            allow_long = True
            allow_short = True
            
            if ma_val_prev is not None and not np.isnan(ma_val_prev):
                allow_long = close_prev > ma_val_prev
                allow_short = close_prev < ma_val_prev
                
            # Long Entry: Supertrend passa da rosso a verde AND close_prev > ma_prev
            if prev_dir == 1 and prev_prev_dir == -1 and (trade_type in ['both', 'long_only']) and allow_long:
                in_position = True
                pos_type = 'long'
                entry_price = current_open
                entry_idx = i
                tp_level = entry_price + tp_multiplier * df['atr'].iloc[i-1]
                
            # Short Entry: Supertrend passa da verde a rosso AND close_prev < ma_prev
            elif prev_dir == -1 and prev_prev_dir == 1 and (trade_type in ['both', 'short_only']) and allow_short:
                in_position = True
                pos_type = 'short'
                entry_price = current_open
                entry_idx = i
                tp_level = entry_price - tp_multiplier * df['atr'].iloc[i-1]
                
    return trades

def format_excel_workbook(filepath):
    """Formatta graficamente il file Excel per renderlo professionale e leggibile."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(filepath)
    
    # Stili professionali
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Blu scuro
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    title_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # Azzurro chiaro
    title_font = Font(name="Calibri", size=14, bold=True, color="1F497D")
    
    regular_font = Font(name="Calibri", size=11)
    bold_font = Font(name="Calibri", size=11, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    double_bottom_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Verde chiaro per profitto
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")   # Rosso chiaro per perdita
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 1. Trova le intestazioni reali cercando la prima riga non vuota
        header_row = 1
        
        # Format Intestazioni colonne
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        ws.row_dimensions[header_row].height = 28
        
        # 2. Formatta i dati
        for row in range(header_row + 1, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
            
            # Leggi il tipo di operazione per capire se evidenziare in verde/rosso i profitti
            # pnl_punti è alla colonna 7, pnl_eur alla colonna 8
            pnl_val = ws.cell(row=row, column=7).value
            pnl_fill = None
            if pnl_val is not None:
                try:
                    pnl_num = float(pnl_val)
                    pnl_fill = green_fill if pnl_num > 0 else red_fill
                except ValueError:
                    pass
            
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = regular_font
                cell.border = thin_border
                
                # Allineamento e formato specifico
                val = cell.value
                
                # Colonne Date (col 2, 4)
                if col in [2, 4]:
                    cell.alignment = Alignment(horizontal="center")
                    # Se ha l'ora, formatta YYYY-MM-DD HH:MM, altrimenti YYYY-MM-DD
                    if isinstance(val, str) and len(val) > 10:
                        cell.number_format = 'yyyy-mm-dd hh:mm'
                    else:
                        cell.number_format = 'yyyy-mm-dd'
                # Colonne Tipo e Motivo Uscita
                elif col in [1, 6]:
                    cell.alignment = Alignment(horizontal="center")
                # Prezzi (col 3, 5 e le colonne OHLC di mercato)
                elif col in [3, 5] or col >= 10:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'
                # PnL Punti e EUR
                elif col in [7, 8]:
                    cell.alignment = Alignment(horizontal="right")
                    if col == 7:
                        cell.number_format = '+#,##0.00;-#,##0.00;0.00'
                    else:
                        cell.number_format = '+#,##0.00" €";-#,##0.00" €";0.00" €"'
                    if pnl_fill:
                        cell.fill = pnl_fill
                # Durata
                elif col == 9:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0'
        
        # 3. Adatta la larghezza delle colonne
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            max_len = 0
            # Scorri le celle per trovare la lunghezza massima
            for row in range(header_row, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            # Imposta larghezza con padding
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    wb.save(filepath)
    print(f"Formattazione completata con successo su '{filepath}'.")

def main():
    print("Avvio della generazione del file Excel...")
    
    # Carica i due dataset
    df_daily = load_data('giornaliero btp.txt')
    df_220m = load_data('btp_220m.txt')
    
    print(f"Dati Daily caricati: {len(df_daily)} righe.")
    print(f"Dati 220m caricati: {len(df_220m)} righe.")
    
    # 1. Calcoli per Daily
    df_daily = calculate_supertrend(df_daily, period=20, multiplier=1.5, atr_type='rma')
    df_daily['sma21'] = df_daily['close'].rolling(window=21).mean()
    
    # 2. Calcoli per 220m
    df_220m = calculate_supertrend(df_220m, period=20, multiplier=1.5, atr_type='rma')
    df_220m['sma21'] = df_220m['close'].rolling(window=21).mean()
    
    # Esecuzione dei backtest
    # Configurazione 1: Con filtro SMA 21 (Strategia principale)
    trades_daily_sma = run_backtest(df_daily, tp_multiplier=999.0, ma_col='sma21')
    trades_220m_sma = run_backtest(df_220m, tp_multiplier=999.0, ma_col='sma21')
    
    # Configurazione 2: Senza filtro SMA 21 (Per confronto)
    trades_daily_nofilter = run_backtest(df_daily, tp_multiplier=999.0, ma_col=None)
    trades_220m_nofilter = run_backtest(df_220m, tp_multiplier=999.0, ma_col=None)
    
    # Conversione in DataFrame
    df_trades_daily_sma = pd.DataFrame(trades_daily_sma)
    df_trades_220m_sma = pd.DataFrame(trades_220m_sma)
    df_trades_daily_nofilter = pd.DataFrame(trades_daily_nofilter)
    df_trades_220m_nofilter = pd.DataFrame(trades_220m_nofilter)
    
    # Rinomina colonne per renderle chiare ed eleganti
    rename_cols = {
        'tipo': 'Tipo',
        'data_ingresso': 'Data Ingresso',
        'prezzo_ingresso': 'Prezzo Ingresso',
        'data_uscita': 'Data Uscita',
        'prezzo_uscita': 'Prezzo Uscita',
        'motivo_uscita': 'Motivo Uscita',
        'pnl_punti': 'PnL (Punti)',
        'pnl_eur': 'PnL (EUR)',
        'durata_barre': 'Durata (Barre)',
        'mercato_open_ingresso': 'Mercato Open (Ingresso)',
        'mercato_high_ingresso': 'Mercato High (Ingresso)',
        'mercato_low_ingresso': 'Mercato Low (Ingresso)',
        'mercato_close_ingresso': 'Mercato Close (Ingresso)',
        'mercato_open_uscita': 'Mercato Open (Uscita)',
        'mercato_high_uscita': 'Mercato High (Uscita)',
        'mercato_low_uscita': 'Mercato Low (Uscita)',
        'mercato_close_uscita': 'Mercato Close (Uscita)'
    }
    
    for df_t in [df_trades_daily_sma, df_trades_220m_sma, df_trades_daily_nofilter, df_trades_220m_nofilter]:
        if not df_t.empty:
            df_t.rename(columns=rename_cols, inplace=True)
            
    # Salva in Excel con fogli multipli
    output_filename = 'btp_trades_report.xlsx'
    
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        if not df_trades_220m_sma.empty:
            df_trades_220m_sma.to_excel(writer, sheet_name='Trades_220m_SMA21', index=False)
        else:
            pd.DataFrame(columns=rename_cols.values()).to_excel(writer, sheet_name='Trades_220m_SMA21', index=False)
            
        if not df_trades_daily_sma.empty:
            df_trades_daily_sma.to_excel(writer, sheet_name='Trades_Daily_SMA21', index=False)
        else:
            pd.DataFrame(columns=rename_cols.values()).to_excel(writer, sheet_name='Trades_Daily_SMA21', index=False)
            
        if not df_trades_220m_nofilter.empty:
            df_trades_220m_nofilter.to_excel(writer, sheet_name='Trades_220m_NoFilter', index=False)
        else:
            pd.DataFrame(columns=rename_cols.values()).to_excel(writer, sheet_name='Trades_220m_NoFilter', index=False)
            
        if not df_trades_daily_nofilter.empty:
            df_trades_daily_nofilter.to_excel(writer, sheet_name='Trades_Daily_NoFilter', index=False)
        else:
            pd.DataFrame(columns=rename_cols.values()).to_excel(writer, sheet_name='Trades_Daily_NoFilter', index=False)
            
    print(f"File Excel grezzo scritto in '{output_filename}'.")
    
    # Formattazione estetica premium del file Excel
    format_excel_workbook(output_filename)
    
    print("\nProcesso completato! Il file Excel premium 'btp_trades_report.xlsx' è pronto.")

if __name__ == '__main__':
    main()
